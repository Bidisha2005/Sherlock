"""
QA Rubric Generator - Automated Excel Sheet Creator
Extracts JDs from job portals and creates structured rubric
"""

import pandas as pd
import re
import json
import time
import os
from datetime import datetime
from typing import List, Dict, Set, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import requests
from bs4 import BeautifulSoup
import googlesearch
import openai  # Optional, for ChatGPT integration

# ================================
# CONFIGURATION
# ================================

class Config:
    # Job Portals (URL patterns for search)
    JOB_PORTALS = {
        'linkedin': 'https://www.linkedin.com/jobs/search/?keywords=QA%20Engineer',
        'indeed': 'https://www.indeed.com/q-QA-Engineer-jobs.html',
        'monster': 'https://www.monster.com/jobs/search/?q=QA-Engineer',
        'dice': 'https://www.dice.com/jobs/q-qa-engineer-jobs',
        'glassdoor': 'https://www.glassdoor.com/Job/qa-engineer-jobs-SRCH_KO0,11.htm'
    }
    
    TARGET_JD_COUNT = 60
    OUTPUT_FILE = 'QA_Rubric.xlsx'
    JD_TEXT_FILE = 'QA_JDs_Extraction.txt'
    
    # Common QA Skills (for reference)
    QA_SKILLS = {
        'Manual Testing': ['Test Cases', 'Test Plans', 'Regression Testing', 'Smoke Testing', 'Sanity Testing'],
        'Automation Testing': ['Selenium', 'Cypress', 'Playwright', 'TestNG', 'JUnit', 'Cucumber', 'Appium'],
        'API Testing': ['Postman', 'RestAssured', 'SOAP UI', 'Swagger', 'JMeter'],
        'Performance Testing': ['JMeter', 'LoadRunner', 'Gatling', 'Artillery'],
        'Security Testing': ['OWASP', 'Burp Suite', 'SonarQube', 'ZAP'],
        'Database Testing': ['SQL', 'MySQL', 'PostgreSQL', 'MongoDB'],
        'Mobile Testing': ['Appium', 'XCUITest', 'Espresso', 'Detox'],
        'Bug Tracking': ['Jira', 'Bugzilla', 'Asana', 'Trello', 'Azure DevOps'],
        'CI/CD': ['Jenkins', 'GitHub Actions', 'GitLab CI', 'CircleCI'],
        'Cloud Testing': ['AWS', 'GCP', 'Azure', 'LambdaTest', 'BrowserStack']
    }
    
    # ChatGPT API Key (optional)
    OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')

# ================================
# JD EXTRACTION MODULE
# ================================

class JDExtractor:
    """Extract Job Descriptions from various portals"""
    
    def __init__(self, target_count=60):
        self.target_count = target_count
        self.extracted_jds = []
        self.driver = None
        
    def setup_driver(self):
        """Setup Chrome driver for web scraping"""
        options = Options()
        options.add_argument('--headless')  # Run in background
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        self.driver = webdriver.Chrome(options=options)
        
    def extract_linkedin_jds(self, count=15):
        """Extract JDs from LinkedIn"""
        jds = []
        try:
            self.driver.get(Config.JOB_PORTALS['linkedin'])
            time.sleep(5)
            
            # Scroll to load more jobs
            for _ in range(3):
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            
            job_cards = self.driver.find_elements(By.CLASS_NAME, 'job-card-container')
            
            for card in job_cards[:count]:
                try:
                    title = card.find_element(By.CLASS_NAME, 'job-card-container__title').text
                    company = card.find_element(By.CLASS_NAME, 'job-card-container__company-name').text
                    
                    # Click to get full description
                    card.click()
                    time.sleep(1)
                    
                    desc_element = WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CLASS_NAME, 'job-details'))
                    )
                    description = desc_element.text
                    
                    jd = {
                        'portal': 'LinkedIn',
                        'title': title,
                        'company': company,
                        'description': description,
                        'skills': self.extract_skills(description)
                    }
                    jds.append(jd)
                    
                except Exception as e:
                    print(f"Error extracting LinkedIn JD: {e}")
                    continue
                    
        except Exception as e:
            print(f"LinkedIn extraction error: {e}")
            
        return jds
    
    def extract_indeed_jds(self, count=15):
        """Extract JDs from Indeed"""
        jds = []
        try:
            self.driver.get(Config.JOB_PORTALS['indeed'])
            time.sleep(5)
            
            job_cards = self.driver.find_elements(By.CLASS_NAME, 'job_seen_beacon')
            
            for card in job_cards[:count]:
                try:
                    title = card.find_element(By.CLASS_NAME, 'jobTitle').text
                    company = card.find_element(By.CLASS_NAME, 'companyName').text
                    
                    # Get description
                    desc_elements = card.find_elements(By.CLASS_NAME, 'job-snippet')
                    description = ' '.join([el.text for el in desc_elements])
                    
                    jd = {
                        'portal': 'Indeed',
                        'title': title,
                        'company': company,
                        'description': description,
                        'skills': self.extract_skills(description)
                    }
                    jds.append(jd)
                    
                except Exception as e:
                    print(f"Error extracting Indeed JD: {e}")
                    continue
                    
        except Exception as e:
            print(f"Indeed extraction error: {e}")
            
        return jds
    
    def extract_skills(self, text: str) -> List[str]:
        """Extract QA-related skills from text using regex patterns"""
        skills_found = []
        
        # Pattern matching for common QA skills
        skill_patterns = {
            'automation': ['selenium', 'cypress', 'playwright', 'testng', 'junit', 'cucumber', 'appium'],
            'manual': ['manual testing', 'test case', 'test plan', 'regression', 'smoke test'],
            'api': ['postman', 'restassured', 'soap', 'swagger', 'api testing'],
            'performance': ['jmeter', 'loadrunner', 'performance', 'load test', 'stress test'],
            'security': ['owasp', 'burp', 'security test', 'penetration'],
            'database': ['sql', 'mysql', 'postgresql', 'database', 'query'],
            'mobile': ['appium', 'xcuitest', 'espresso', 'mobile test'],
            'bug': ['jira', 'bugzilla', 'defect', 'bug tracking'],
            'ci_cd': ['jenkins', 'github actions', 'gitlab ci', 'circleci'],
            'cloud': ['aws', 'gcp', 'azure', 'lambda', 'browserstack']
        }
        
        text_lower = text.lower()
        
        for category, keywords in skill_patterns.items():
            for keyword in keywords:
                if keyword in text_lower:
                    skills_found.append(keyword)
        
        return list(set(skills_found))  # Remove duplicates
    
    def extract_all_jds(self) -> List[Dict]:
        """Extract JDs from all portals"""
        self.setup_driver()
        
        all_jds = []
        
        # Extract from each portal
        extraction_methods = [
            self.extract_linkedin_jds,
            self.extract_indeed_jds
        ]
        
        for method in extraction_methods:
            try:
                jds = method(self.target_count // len(extraction_methods))
                all_jds.extend(jds)
                print(f"Extracted {len(jds)} JDs from {method.__name__}")
            except Exception as e:
                print(f"Error in {method.__name__}: {e}")
        
        self.driver.quit()
        
        # Trim to target count
        self.extracted_jds = all_jds[:self.target_count]
        return self.extracted_jds
    
    def save_jds_to_file(self, filename='QA_JDs_Extraction.txt'):
        """Save extracted JDs to a text file"""
        with open(filename, 'w', encoding='utf-8') as f:
            for i, jd in enumerate(self.extracted_jds, 1):
                f.write(f"{'='*60}\n")
                f.write(f"--- JD #{i} ---\n")
                f.write(f"Portal: {jd.get('portal', 'Unknown')}\n")
                f.write(f"Job Title: {jd.get('title', 'N/A')}\n")
                f.write(f"Company: {jd.get('company', 'N/A')}\n")
                f.write(f"Date: {datetime.now().strftime('%Y-%m-%d')}\n\n")
                f.write("Required Skills & Keywords:\n")
                for skill in jd.get('skills', []):
                    f.write(f"- {skill}\n")
                f.write(f"\nJob Description:\n{jd.get('description', 'N/A')[:500]}...\n")
                f.write(f"{'='*60}\n\n")

# ================================
# RUBRIC GENERATOR MODULE
# ================================

class RubricGenerator:
    """Generate QA Rubric from extracted JDs and enrichment"""
    
    def __init__(self):
        self.primary_skills = {}
        self.rubric_data = []
        self.jds = []
        
    def load_jds(self, jds: List[Dict]):
        """Load extracted JDs"""
        self.jds = jds
        
    def identify_primary_skills(self) -> Dict[str, Set[str]]:
        """Identify primary skills and their associated terms from JDs"""
        skill_mapping = {}
        
        # Initialize with known QA skills
        for primary, associated in Config.QA_SKILLS.items():
            skill_mapping[primary] = set(associated)
        
        # Enrich from JDs
        for jd in self.jds:
            skills = jd.get('skills', [])
            for skill in skills:
                # Try to map to a primary skill
                for primary, associated in Config.QA_SKILLS.items():
                    if skill in associated:
                        skill_mapping[primary].add(skill)
                        break
                else:
                    # If skill not found, add to a "Other" category
                    if 'Other QA Skills' not in skill_mapping:
                        skill_mapping['Other QA Skills'] = set()
                    skill_mapping['Other QA Skills'].add(skill)
        
        self.primary_skills = {k: list(v) for k, v in skill_mapping.items()}
        return self.primary_skills
    
    def enrich_with_chatgpt(self, primary_skill: str) -> Dict:
        """
        Enrich skill information using ChatGPT API
        """
        if not Config.OPENAI_API_KEY:
            return self._mock_enrichment(primary_skill)
        
        try:
            openai.api_key = Config.OPENAI_API_KEY
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a QA expert. Provide structured information about QA skills."},
                    {"role": "user", "content": f"""
                    For the QA primary skill '{primary_skill}', provide:
                    1. Associated skills (tools, frameworks) - comma separated
                    2. Associated terms (concepts, features) - comma separated  
                    3. Areas of testing - comma separated
                    4. Associated languages - comma separated
                    
                    Format: JSON with keys: 'associated_skills', 'associated_terms', 'areas_of_testing', 'associated_languages'
                    
                    Only include verified, industry-standard items.
                    """}
                ],
                temperature=0.3
            )
            
            result = json.loads(response.choices[0].message['content'])
            return result
            
        except Exception as e:
            print(f"ChatGPT API error: {e}")
            return self._mock_enrichment(primary_skill)
    
    def _mock_enrichment(self, primary_skill: str) -> Dict:
        """Mock enrichment when API is not available"""
        enrichment_map = {
            'Manual Testing': {
                'associated_skills': 'TestRail, QTest, PractiTest',
                'associated_terms': 'Test Cases, Test Plans, Regression, Smoke Testing',
                'areas_of_testing': 'Functional Testing, Exploratory Testing, Usability Testing',
                'associated_languages': 'English, Business Language'
            },
            'Automation Testing': {
                'associated_skills': 'Selenium, Cypress, Playwright, TestNG, JUnit, Cucumber, Appium',
                'associated_terms': 'Page Object Model, Test Framework, Test Scripts, Execution Reports',
                'areas_of_testing': 'Web Automation, Mobile Automation, Desktop Automation',
                'associated_languages': 'Java, Python, JavaScript, C#'
            },
            'API Testing': {
                'associated_skills': 'Postman, RestAssured, SOAP UI, Swagger',
                'associated_terms': 'Endpoints, Requests, Responses, Authentication, JSON/XML Validation',
                'areas_of_testing': 'REST API, SOAP API, GraphQL, Microservices',
                'associated_languages': 'Java, Python, JavaScript'
            },
            'Performance Testing': {
                'associated_skills': 'JMeter, LoadRunner, Gatling, Artillery',
                'associated_terms': 'Load Testing, Stress Testing, Endurance, Spike Testing',
                'areas_of_testing': 'Web Performance, API Performance, Mobile Performance',
                'associated_languages': 'Java, Python, JavaScript'
            },
            'Security Testing': {
                'associated_skills': 'OWASP ZAP, Burp Suite, SonarQube, Acunetix',
                'associated_terms': 'Vulnerability Scanning, Penetration Testing, Encryption, Authentication',
                'areas_of_testing': 'Web Security, API Security, Mobile Security',
                'associated_languages': 'Python, Ruby, Java'
            },
            'Database Testing': {
                'associated_skills': 'SQL, MySQL, PostgreSQL, MongoDB, Oracle',
                'associated_terms': 'Queries, Stored Procedures, Data Integrity, Joins, Transactions',
                'areas_of_testing': 'Data Validation, ETL Testing, Data Warehouse Testing',
                'associated_languages': 'SQL, PL/SQL, T-SQL'
            },
            'Mobile Testing': {
                'associated_skills': 'Appium, XCUITest, Espresso, Detox, BrowserStack',
                'associated_terms': 'Device Fragmentation, OS Versions, Functional Testing, Payment Testing',
                'areas_of_testing': 'iOS Testing, Android Testing, Cross-platform Testing',
                'associated_languages': 'Java, Kotlin, Swift, JavaScript'
            },
            'Bug Tracking': {
                'associated_skills': 'Jira, Bugzilla, Asana, Trello, Azure DevOps',
                'associated_terms': 'Defect Lifecycle, Bug Reports, Severity, Priority, Workflow',
                'areas_of_testing': 'Bug Management, Issue Tracking, Project Management',
                'associated_languages': 'English'
            },
            'CI/CD Testing': {
                'associated_skills': 'Jenkins, GitHub Actions, GitLab CI, CircleCI, Travis CI',
                'associated_terms': 'Continuous Testing, Pipeline Integration, Build Verification',
                'areas_of_testing': 'DevOps Testing, Pipeline Testing, Automated Builds',
                'associated_languages': 'Groovy, YAML, Python'
            },
            'Cloud Testing': {
                'associated_skills': 'AWS, GCP, Azure, LambdaTest, SauceLabs',
                'associated_terms': 'Cloud Infrastructure, Scaling, Load Balancing, Disaster Recovery',
                'areas_of_testing': 'Cloud Migration, Cloud-native Testing, Hybrid Cloud',
                'associated_languages': 'Python, Java, Go'
            }
        }
        
        # Default for unknown skills
        default_enrichment = {
            'associated_skills': 'Varied tools and frameworks',
            'associated_terms': 'Concepts and methodologies',
            'areas_of_testing': 'General testing areas',
            'associated_languages': 'Multiple languages'
        }
        
        return enrichment_map.get(primary_skill, default_enrichment)
    
    def build_rubric_data(self) -> List[Dict]:
        """Build the complete rubric data"""
        self.identify_primary_skills()
        
        for primary_skill, associated_skills in self.primary_skills.items():
            enrichment = self.enrich_with_chatgpt(primary_skill)
            
            row = {
                'Profile': 'QA (Quality Assurance)',
                'Primary Skill': primary_skill,
                'Associated Skills': ', '.join(associated_skills),
                'Associated Terms': enrichment.get('associated_terms', ''),
                'Area of Testing': enrichment.get('areas_of_testing', ''),
                'Associated Languages': enrichment.get('associated_languages', '')
            }
            
            self.rubric_data.append(row)
        
        return self.rubric_data

# ================================
# EXCEL GENERATOR MODULE
# ================================

class ExcelGenerator:
    """Generate and style the Excel rubric file"""
    
    def __init__(self, data: List[Dict], filename='QA_Rubric.xlsx'):
        self.data = data
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = 'QA Rubric'
        
    def generate_excel(self):
        """Generate styled Excel file"""
        # Define headers
        headers = ['Profile', 'Primary Skill', 'Associated Skills', 
                   'Associated Terms', 'Area of Testing', 'Associated Languages']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_idx, row_data in enumerate(self.data, 2):
            self.ws.cell(row=row_idx, column=1, value=row_data['Profile'])
            self.ws.cell(row=row_idx, column=2, value=row_data['Primary Skill'])
            self.ws.cell(row=row_idx, column=3, value=row_data['Associated Skills'])
            self.ws.cell(row=row_idx, column=4, value=row_data['Associated Terms'])
            self.ws.cell(row=row_idx, column=5, value=row_data['Area of Testing'])
            self.ws.cell(row=row_idx, column=6, value=row_data['Associated Languages'])
        
        # Auto-adjust column widths
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.ws.iter_rows(min_row=1, max_row=len(self.data)+1, 
                                     min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Save the file
        self.wb.save(self.filename)
        print(f"✅ Excel file saved as: {self.filename}")
        
    def create_pivot_analysis(self):
        """Create an additional sheet with analysis"""
        ws_pivot = self.wb.create_sheet('Skill Analysis')
        
        # Count occurrences of each primary skill
        skill_counts = {}
        for row in self.data:
            skill = row['Primary Skill']
            skill_counts[skill] = skill_counts.get(skill, 0) + 1
        
        # Write skill analysis
        ws_pivot.cell(row=1, column=1, value='Primary Skill')
        ws_pivot.cell(row=1, column=2, value='Frequency in JDs')
        
        for idx, (skill, count) in enumerate(sorted(skill_counts.items(), 
                                                     key=lambda x: x[1], reverse=True), 2):
            ws_pivot.cell(row=idx, column=1, value=skill)
            ws_pivot.cell(row=idx, column=2, value=count)
        
        # Save again with pivot sheet
        self.wb.save(self.filename)

# ================================
# MAIN EXECUTION
# ================================

def main():
    """Main execution function"""
    print("🚀 Starting QA Rubric Automation...")
    print("="*60)
    
    # Step 1: Extract JDs
    print("📋 Extracting Job Descriptions from 5 portals...")
    extractor = JDExtractor(target_count=Config.TARGET_JD_COUNT)
    jds = extractor.extract_all_jds()
    print(f"✅ Extracted {len(jds)} Job Descriptions")
    
    # Step 2: Save JDs to text file
    extractor.save_jds_to_file(Config.JD_TEXT_FILE)
    print(f"💾 JDs saved to: {Config.JD_TEXT_FILE}")
    
    # Step 3: Generate Rubric
    print("📊 Generating QA Rubric...")
    generator = RubricGenerator()
    generator.load_jds(jds)
    rubric_data = generator.build_rubric_data()
    print(f"✅ Generated {len(rubric_data)} primary skills")
    
    # Step 4: Create Excel
    print("📁 Creating Excel file...")
    excel_gen = ExcelGenerator(rubric_data, Config.OUTPUT_FILE)
    excel_gen.generate_excel()
    excel_gen.create_pivot_analysis()
    
    print("="*60)
    print("🎉 Automation Complete!")
    print(f"📄 Output file: {Config.OUTPUT_FILE}")
    print(f"📄 JD text file: {Config.JD_TEXT_FILE}")
    print("\n📌 Summary:")
    print(f"   - Total JDs extracted: {len(jds)}")
    print(f"   - Primary skills identified: {len(rubric_data)}")
    print(f"   - Total associated skills: {sum(len(row['Associated Skills'].split(', ')) for row in rubric_data)}")

if __name__ == "__main__":
    main()